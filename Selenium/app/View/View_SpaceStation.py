from lib2to3.pgen2 import driver
from logging import setLogRecordFactory
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time


class SpaceStation:
    driver = None
    start_time = time.time()
    """

    """

    email_field_xpath = '//input[contains(@id, "email")]'
    password_field_xpath = '//input[contains(@id, "password")]'
    signIn_btn_xpath = '//*[@id="root"]/div/header/div/div/div/div[2]/button'

    def __init__(self, driver) -> None:
        self.driver = driver
        self.email_field = WebDriverWait(self.driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, self.email_field_xpath)))

    def insert_email(self, value):
        self.email_field.send_keys(value)

    def insert_password(self, value):
        self.password_field = WebDriverWait(self.driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, self.password_field_xpath)))
        self.password_field.send_keys(value)

    def click_signIn(self):
        self.signIn_btn = WebDriverWait(self.driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, self.signIn_btn_xpath)))
        self.signIn_btn.click()

    def reset_timer(self):
        self.start_time = time.time()

    def save_results(self, feature, scenario, filename):
        import openpyxl
        import os

        end_time = time.time()

        date = time.strftime("%Y-%m-%d", time.gmtime())

        path = f"{filename}_{date}.xlsx"

        book = openpyxl.Workbook()
        if os.path.isfile(path):
            book = openpyxl.load_workbook(path)

        isSheet = False

        for sh in book:
            if sh.title == feature:
                sheet = book[feature]
                isSheet = True

        if isSheet == False:
            book.create_sheet(feature)

        if book.get_sheet_names()[0] == "Sheet":
            sheet_to_del = book['Sheet']
            book.remove(sheet_to_del)

        sheet = book[feature]

        diff = end_time - self.start_time
        sheet.append([scenario])
        sheet.append(["StartTime", "EndTime", "Difference"])
        sheet.append([self.start_time, end_time, diff])
        sheet.append([""])

        book.save(f"{filename}_{date}.xlsx")
