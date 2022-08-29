import os
from re import I
import openpyxl
import json
import sys

data = {}
language = sys.argv[1]
test = sys.argv[2]

with open(f"SpaceStation_{language}_MAX_{test}.postman_test_run.json", "r") as read_file:
    data_max = json.load(read_file)

results_max = data_max["results"]

with open(f"SpaceStation_{language}_MED_{test}.postman_test_run.json", "r") as read_file:
    data_med = json.load(read_file)

results_med = data_med["results"]

with open(f"SpaceStation_{language}_MIN_{test}.postman_test_run.json", "r") as read_file:
    data_min = json.load(read_file)

results_min = data_min["results"]

path = f"{language}_Postman.xlsx"

book = openpyxl.Workbook()
if os.path.isfile(path):
    book = openpyxl.load_workbook(path)

sum_max = 0.0
sum_med = 0.0
sum_min = 0.0

for i in range(len(results_max)):
    sheetName = "Data"

    isSheet = False

    for sh in book:
        if sh.title == sheetName:
            isSheet = True

    if isSheet == False:
        book.create_sheet(sheetName)

    sheet = book[sheetName]

    time_max = float(results_max[i]['times'][0]) * 0.001
    sum_max += time_max
    time_med = float(results_med[i]['times'][0]) * 0.001
    sum_med += time_med
    time_min = float(results_min[i]['times'][0]) * 0.001
    sum_min += time_min

    sheet.append([results_max[i]['name']])
    sheet.append(["", "", "Time"])
    sheet.append(["", "MAX", time_max])
    sheet.append(["", "MED", time_med])
    sheet.append(["", "MIN", time_min])
    sheet.append([""])

if book.sheetnames[0] == "Sheet":
    sheet_to_del = book['Sheet']
    book.remove(sheet_to_del)

book.create_sheet("Summary")
sheet = book["Summary"]

sheet.append(["", "Sum"])
sheet.append(["MAX", sum_max])
sheet.append(["MED", sum_med])
sheet.append(["MIN", sum_min])

book.save(f"{language}_Postman_{test}_a.xlsx")
