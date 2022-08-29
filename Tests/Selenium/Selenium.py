import os
import sys
import openpyxl
import time

language = sys.argv[1]
test = sys.argv[2]

date = time.strftime("%Y-%m-%d", time.gmtime())

path = f""

data_min = openpyxl.Workbook()
data_med = openpyxl.Workbook()
data_max = openpyxl.Workbook()

data_max = openpyxl.load_workbook(
    filename=f"{language}_MAX_{test}_{date}.xlsx")
data_med = openpyxl.load_workbook(
    filename=f"{language}_MED_{test}_{date}.xlsx")
data_min = openpyxl.load_workbook(
    filename=f"{language}_MIN_{test}_{date}.xlsx")


book = openpyxl.Workbook()
if os.path.isfile(path):
    book = openpyxl.load_workbook(path)

sum_max = 0.0
sum_med = 0.0
sum_min = 0.0

for sh in data_max:
    book.create_sheet(sh.title)
    sheet = book[sh.title]

    if book.get_sheet_names()[0] == "Sheet":
        sheet_to_del = book['Sheet']
        book.remove(sheet_to_del)

    sheet_max = data_max[sh.title]
    sheet_med = data_med[sh.title]
    sheet_min = data_min[sh.title]

    if sh.title == "Login":
        sheet.append(["Correct user login"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A3"].value,
                      sheet_max["B3"].value, sheet_max["C3"].value])
        sheet.append(["MED", sheet_med["A3"].value,
                      sheet_med["B3"].value, sheet_med["C3"].value])
        sheet.append(["MIN", sheet_min["A3"].value,
                      sheet_min["B3"].value, sheet_min["C3"].value])
        sum_max += sheet_max["C3"].value
        sum_med += sheet_med["C3"].value
        sum_min += sheet_min["C3"].value

    elif sh.title == "Missions":
        sheet.append(["Show missions crew"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A3"].value,
                      sheet_max["B3"].value, sheet_max["C3"].value])
        sheet.append(["MED", sheet_med["A3"].value,
                      sheet_med["B3"].value, sheet_med["C3"].value])
        sheet.append(["MIN", sheet_min["A3"].value,
                      sheet_min["B3"].value, sheet_min["C3"].value])
        sum_max += sheet_max["C3"].value
        sum_med += sheet_med["C3"].value
        sum_min += sheet_min["C3"].value

        sheet.append([""])
        sheet.append(["Add missions"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A7"].value,
                      sheet_max["B7"].value, sheet_max["C7"].value])
        sheet.append(["MED", sheet_med["A7"].value,
                      sheet_med["B7"].value, sheet_med["C7"].value])
        sheet.append(["MIN", sheet_min["A7"].value,
                      sheet_min["B7"].value, sheet_min["C7"].value])
        sum_max += sheet_max["C7"].value
        sum_med += sheet_med["C7"].value
        sum_min += sheet_min["C7"].value

        sheet.append([""])
        sheet.append(["Add missions crew"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A11"].value,
                      sheet_max["B11"].value, sheet_max["C11"].value])
        sheet.append(["MED", sheet_med["A11"].value,
                      sheet_med["B11"].value, sheet_med["C11"].value])
        sheet.append(["MIN", sheet_min["A11"].value,
                      sheet_min["B11"].value, sheet_min["C11"].value])
        sum_max += sheet_max["C11"].value
        sum_med += sheet_med["C11"].value
        sum_min += sheet_min["C11"].value

    elif sh.title == "Reports":
        sheet.append(["Add Report"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A3"].value,
                      sheet_max["B3"].value, sheet_max["C3"].value])
        sheet.append(["MED", sheet_med["A3"].value,
                      sheet_med["B3"].value, sheet_med["C3"].value])
        sheet.append(["MIN", sheet_min["A3"].value,
                      sheet_min["B3"].value, sheet_min["C3"].value])
        sum_max += sheet_max["C3"].value
        sum_med += sheet_med["C3"].value
        sum_min += sheet_min["C3"].value

        sheet.append([""])
        sheet.append(["Assign Report"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A7"].value,
                      sheet_max["B7"].value, sheet_max["C7"].value])
        sheet.append(["MED", sheet_med["A7"].value,
                      sheet_med["B7"].value, sheet_med["C7"].value])
        sheet.append(["MIN", sheet_min["A7"].value,
                      sheet_min["B7"].value, sheet_min["C7"].value])
        sum_max += sheet_max["C7"].value
        sum_med += sheet_med["C7"].value
        sum_min += sheet_min["C7"].value

        sheet.append([""])
        sheet.append(["Change report status to in progress"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A11"].value,
                      sheet_max["B11"].value, sheet_max["C11"].value])
        sheet.append(["MED", sheet_med["A11"].value,
                      sheet_med["B11"].value, sheet_med["C11"].value])
        sheet.append(["MIN", sheet_min["A11"].value,
                      sheet_min["B11"].value, sheet_min["C11"].value])
        sum_max += sheet_max["C11"].value
        sum_med += sheet_med["C11"].value
        sum_min += sheet_min["C11"].value

        sheet.append([""])
        sheet.append(["Change report status to in finished"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A15"].value,
                      sheet_max["B15"].value, sheet_max["C15"].value])
        sheet.append(["MED", sheet_med["A15"].value,
                      sheet_med["B15"].value, sheet_med["C15"].value])
        sheet.append(["MIN", sheet_min["A15"].value,
                      sheet_min["B15"].value, sheet_min["C15"].value])
        sum_max += sheet_max["C15"].value
        sum_med += sheet_med["C15"].value
        sum_min += sheet_min["C15"].value

        sheet.append([""])
        sheet.append(["Add and cancel report"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A19"].value,
                      sheet_max["B19"].value, sheet_max["C19"].value])
        sheet.append(["MED", sheet_med["A19"].value,
                      sheet_med["B19"].value, sheet_med["C19"].value])
        sheet.append(["MIN", sheet_min["A19"].value,
                      sheet_min["B19"].value, sheet_min["C19"].value])
        sum_max += sheet_max["C19"].value
        sum_med += sheet_med["C19"].value
        sum_min += sheet_min["C19"].value

    elif sh.title == "Users":
        sheet.append(["Add user manager"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A3"].value,
                      sheet_max["B3"].value, sheet_max["C3"].value])
        sheet.append(["MED", sheet_med["A3"].value,
                      sheet_med["B3"].value, sheet_med["C3"].value])
        sheet.append(["MIN", sheet_min["A3"].value,
                      sheet_min["B3"].value, sheet_min["C3"].value])
        sum_max += sheet_max["C3"].value
        sum_med += sheet_med["C3"].value
        sum_min += sheet_min["C3"].value

        sheet.append([""])
        sheet.append(["Delete user"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A7"].value,
                      sheet_max["B7"].value, sheet_max["C7"].value])
        sheet.append(["MED", sheet_med["A7"].value,
                      sheet_med["B7"].value, sheet_med["C7"].value])
        sheet.append(["MIN", sheet_min["A7"].value,
                      sheet_min["B7"].value, sheet_min["C7"].value])
        sum_max += sheet_max["C7"].value
        sum_med += sheet_med["C7"].value
        sum_min += sheet_min["C7"].value

        sheet.append([""])
        sheet.append(["Add user not manager"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A11"].value,
                      sheet_max["B11"].value, sheet_max["C11"].value])
        sheet.append(["MED", sheet_med["A11"].value,
                      sheet_med["B11"].value, sheet_med["C11"].value])
        sheet.append(["MIN", sheet_min["A11"].value,
                      sheet_min["B11"].value, sheet_min["C11"].value])
        sum_max += sheet_max["C11"].value
        sum_med += sheet_med["C11"].value
        sum_min += sheet_min["C11"].value

        sheet.append([""])
        sheet.append(["Login as not manager"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A15"].value,
                      sheet_max["B15"].value, sheet_max["C15"].value])
        sheet.append(["MED", sheet_med["A15"].value,
                      sheet_med["B15"].value, sheet_med["C15"].value])
        sheet.append(["MIN", sheet_min["A15"].value,
                      sheet_min["B15"].value, sheet_min["C15"].value])
        sum_max += sheet_max["C15"].value
        sum_med += sheet_med["C15"].value
        sum_min += sheet_min["C15"].value

        sheet.append([""])
        sheet.append(["Edit user"])
        sheet.append(["", "StartTime", "EndTime", "Difference"])
        sheet.append(["MAX", sheet_max["A19"].value,
                      sheet_max["B19"].value, sheet_max["C19"].value])
        sheet.append(["MED", sheet_med["A19"].value,
                      sheet_med["B19"].value, sheet_med["C19"].value])
        sheet.append(["MIN", sheet_min["A19"].value,
                      sheet_min["B19"].value, sheet_min["C19"].value])
        sum_max += sheet_max["C19"].value
        sum_med += sheet_med["C19"].value
        sum_min += sheet_min["C19"].value

book.create_sheet("Summary")
sheet = book["Summary"]
sheet.append(["", "Sum"])
sheet.append(["MAX", sum_max])
sheet.append(["MED", sum_med])
sheet.append(["MIN", sum_min])

book.save(f"Selenium_{language}_{test}.xlsx")
