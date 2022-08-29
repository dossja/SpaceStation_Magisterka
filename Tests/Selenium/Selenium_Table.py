import os
from re import I
import re
import openpyxl
import json
import sys
import pandas as pd

data = {}
language = sys.argv[1]

data_Writer = pd.ExcelWriter(f"Selenium_{language}.xlsx")


def iter_rows(sheet1, sheet2, sheet3, sheet4, sheet5, rows, resources):
    titles = []
    values = [[], [], [], [], []]
    for i in range(rows):
        titles.append(sheet1[f"A{i*6 + 1}"].value)
        values[0].append(sheet1[f"D{i*6+resources+2}"].value)
        values[1].append(sheet2[f"D{i*6+resources+2}"].value)
        values[2].append(sheet3[f"D{i*6+resources+2}"].value)
        values[3].append(sheet4[f"D{i*6+resources+2}"].value)
        values[4].append(sheet5[f"D{i*6+resources+2}"].value)
    return titles, values


data1 = openpyxl.load_workbook(
    filename=f"Selenium_{language}_1.xlsx")
data2 = openpyxl.load_workbook(
    filename=f"Selenium_{language}_2.xlsx")
data3 = openpyxl.load_workbook(
    filename=f"Selenium_{language}_3.xlsx")
data4 = openpyxl.load_workbook(
    filename=f"Selenium_{language}_4.xlsx")
data5 = openpyxl.load_workbook(
    filename=f"Selenium_{language}_5.xlsx")


def load_data(resources):
    if resources == "MAX":
        res = 1
    elif resources == "MED":
        res = 2
    else:
        res = 3
    data = pd.DataFrame()
    data_tmp = pd.DataFrame()

    total_titles = []
    total_values = []

    for sh in data1:
        sheet1 = data1[sh.title]
        sheet2 = data2[sh.title]
        sheet3 = data3[sh.title]
        sheet4 = data4[sh.title]
        sheet5 = data5[sh.title]
        rows = 0
        if sh.title == "Login":
            rows = 1
        elif sh.title == "Missions":
            rows = 3
        elif sh.title == "Summary":
            break
        else:
            rows = 5
        titles, values = iter_rows(
            sheet1, sheet2, sheet3, sheet4, sheet5, rows, res)
        for title in titles:
            total_titles.append(title)
        for value in values:
            total_values.append(value)

    total_titles.append("Summary")
    column = [[], [], [], [], []]
    sum = [0, 0, 0, 0, 0]
    for i in range(4):
        if len(total_values[i*5]):
            for id in range(len(total_values[i*5])):
                column[0].append(total_values[i*5][id])
                column[1].append(total_values[i*5+1][id])
                column[2].append(total_values[i*5+2][id])
                column[3].append(total_values[i*5+3][id])
                column[4].append(total_values[i*5+4][id])
                sum[0] += total_values[i*5][id]
                sum[1] += total_values[i*5+1][id]
                sum[2] += total_values[i*5+2][id]
                sum[3] += total_values[i*5+3][id]
                sum[4] += total_values[i*5+4][id]
        else:
            column[0].append(total_values[i*5])
            column[1].append(total_values[i*5+1])
            column[2].append(total_values[i*5+2])
            column[3].append(total_values[i*5+3])
            column[4].append(total_values[i*5+4])
            sum[0] += total_values[i*5]
            sum[1] += total_values[i*5+1]
            sum[2] += total_values[i*5+2]
            sum[3] += total_values[i*5+3]
            sum[4] += total_values[i*5+4]

    column[0].append(sum[0])
    column[1].append(sum[1])
    column[2].append(sum[2])
    column[3].append(sum[3])
    column[4].append(sum[4])

    data["Labels"] = total_titles
    data_tmp["Iter 1"] = column[0]
    data_tmp["Iter 2"] = column[1]
    data_tmp["Iter 3"] = column[2]
    data_tmp["Iter 4"] = column[3]
    data_tmp["Iter 5"] = column[4]

    data["Average"] = data_tmp.mean(axis=1)
    data["Minimum"] = data_tmp.min(axis=1)
    data["Maximum"] = data_tmp.max(axis=1)
    data["St.dev"] = data_tmp.std(axis=1)

    data.to_excel(data_Writer, sheet_name=resources)


load_data("MIN")
load_data("MED")
load_data("MAX")

data_Writer.save()
data_Writer.close()
